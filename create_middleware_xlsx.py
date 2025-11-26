#!/usr/bin/env python3
"""
Create Middleware Testing XLSX
Tracks test results for each CONFIG_SYSTEM_DEVICE module
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

# Create new workbook
wb = Workbook()
wb.remove(wb.active)

# Define colors and styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
pending_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
category_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
category_font = Font(bold=True, color="1F4E78")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# SHEET 1: Middleware Test Results
# ============================================================================

ws_results = wb.create_sheet("Middleware Test Results", 0)

# Header
headers = ["No", "Module Name", "Test Script", "Test 1", "Test 2", "Test 3", "Test 4", "Test 5", "Total Pass", "Pass Rate", "Status", "Notes"]
ws_results.append(headers)

for cell in ws_results[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Middleware modules with their test scripts
modules = [
    # Automation Control (5)
    (1, "AutomationLogic.py", "test_automation_logic.py"),
    (2, "AutomationSchedule.py", "test_automation_schedule.py"),
    (3, "AutomationValue.py", "test_automation_value.py"),
    (4, "AutomationUnified.py", "test_automation_unified.py"),
    (5, "AutomationVoice.py", "test_automation_voice.py"),

    # System Configuration (2)
    (6, "Settings.py", "test_settings.py"),
    (7, "LibraryConfig.py", "test_library_config.py"),

    # Device Management (2)
    (8, "DeviceConfig.py", "test_device_config.py"),
    (9, "RemapPayload.py", "test_remap_payload.py"),

    # Infrastructure (3)
    (10, "BrokerTemplateManager.py", "test_broker_template_manager.py"),
    (11, "ErrorLogger.py", "test_error_logger.py"),
    (12, "snmp_handler.py", "test_snmp_handler.py"),

    # VPN Services (3)
    (13, "ikev2_service.py", "test_ikev2_service.py"),
    (14, "openvpn_service.py", "test_openvpn_service.py"),
    (15, "wireguard_service.py", "test_wireguard_service.py"),

    # Input Control (1)
    (16, "Button.py", "test_button_control.py"),
]

row_num = 2
for no, module, script in modules:
    ws_results[f"A{row_num}"] = no
    ws_results[f"B{row_num}"] = module
    ws_results[f"C{row_num}"] = script

    # Tests 1-5 (Pending by default)
    for col in ["D", "E", "F", "G", "H"]:
        ws_results[f"{col}{row_num}"] = "Pending"
        ws_results[f"{col}{row_num}"].fill = pending_fill
        ws_results[f"{col}{row_num}"].alignment = Alignment(horizontal="center")
        ws_results[f"{col}{row_num}"].border = border

    # Total Pass
    ws_results[f"I{row_num}"] = f'=COUNTIF(D{row_num}:H{row_num},"PASS")'
    ws_results[f"I{row_num}"].alignment = Alignment(horizontal="center")
    ws_results[f"I{row_num}"].border = border

    # Pass Rate
    ws_results[f"J{row_num}"] = f'=IF(I{row_num}=0,"0%",TEXT(I{row_num}/5,"0%"))'
    ws_results[f"J{row_num}"].alignment = Alignment(horizontal="center")
    ws_results[f"J{row_num}"].border = border

    # Status
    ws_results[f"K{row_num}"] = "Pending"
    ws_results[f"K{row_num}"].fill = pending_fill
    ws_results[f"K{row_num}"].alignment = Alignment(horizontal="center")
    ws_results[f"K{row_num}"].border = border

    # Notes
    ws_results[f"L{row_num}"].border = border

    row_num += 1

# Set column widths
ws_results.column_dimensions['A'].width = 5
ws_results.column_dimensions['B'].width = 25
ws_results.column_dimensions['C'].width = 30
ws_results.column_dimensions['D'].width = 10
ws_results.column_dimensions['E'].width = 10
ws_results.column_dimensions['F'].width = 10
ws_results.column_dimensions['G'].width = 10
ws_results.column_dimensions['H'].width = 10
ws_results.column_dimensions['I'].width = 12
ws_results.column_dimensions['J'].width = 12
ws_results.column_dimensions['K'].width = 12
ws_results.column_dimensions['L'].width = 30

# ============================================================================
# SHEET 2: Category Summary
# ============================================================================

ws_summary = wb.create_sheet("Summary by Category", 1)

ws_summary.append(["Category", "Module Count", "Total Tests", "Tests Passed", "Pass Rate", "Status"])

for cell in ws_summary[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

categories = [
    ("Automation Control", 5, 25),
    ("System Configuration", 2, 10),
    ("Device Management", 2, 10),
    ("Infrastructure", 3, 15),
    ("VPN Services", 3, 15),
    ("Input Control", 1, 5),
    ("TOTAL", 16, 80),
]

row_num = 2
for cat, module_count, total_tests in categories:
    ws_summary[f"A{row_num}"] = cat
    ws_summary[f"B{row_num}"] = module_count
    ws_summary[f"C{row_num}"] = total_tests
    ws_summary[f"D{row_num}"] = "0"  # Will be updated with actual results
    ws_summary[f"E{row_num}"] = "0%"
    ws_summary[f"F{row_num}"] = "Pending"

    if cat == "TOTAL":
        ws_summary[f"A{row_num}"].fill = category_fill
        ws_summary[f"A{row_num}"].font = category_font

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws_summary[f"{col}{row_num}"].border = border
        ws_summary[f"{col}{row_num}"].alignment = Alignment(horizontal="center")

    row_num += 1

ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 15
ws_summary.column_dimensions['C'].width = 15
ws_summary.column_dimensions['D'].width = 15
ws_summary.column_dimensions['E'].width = 15
ws_summary.column_dimensions['F'].width = 15

# ============================================================================
# SHEET 3: Test Details
# ============================================================================

ws_details = wb.create_sheet("Test Details", 2)

ws_details.append(["Module", "Test Name", "MQTT Topics", "Operation Type", "Expected Behavior", "Actual Result", "Response Time (ms)", "Status"])

for cell in ws_details[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

test_details = [
    # Automation Logic
    ("AutomationLogic.py", "Get All Rules", "command_control_logic → response_control_logic", "READ", "Returns all logic rules", "", "", "Pending"),
    ("AutomationLogic.py", "Add New Rule", "command_control_logic → response_control_logic", "CREATE", "Creates new rule successfully", "", "", "Pending"),
    ("AutomationLogic.py", "Update Rule", "command_control_logic → response_control_logic", "UPDATE", "Updates rule with new data", "", "", "Pending"),
    ("AutomationLogic.py", "Enable Logging", "command_control_logic → response_control_logic", "ACTION", "Enables device topic logging", "", "", "Pending"),
    ("AutomationLogic.py", "Delete Rule", "command_control_logic → response_control_logic", "DELETE", "Deletes rule successfully", "", "", "Pending"),

    # Add more tests as needed...
]

row_num = 2
for module, test_name, topics, op_type, expected, actual, response_time, status in test_details:
    ws_details[f"A{row_num}"] = module
    ws_details[f"B{row_num}"] = test_name
    ws_details[f"C{row_num}"] = topics
    ws_details[f"D{row_num}"] = op_type
    ws_details[f"E{row_num}"] = expected
    ws_details[f"F{row_num}"] = actual
    ws_details[f"G{row_num}"] = response_time
    ws_details[f"H{row_num}"] = status

    if status == "Pending":
        ws_details[f"H{row_num}"].fill = pending_fill

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws_details[f"{col}{row_num}"].border = border
        ws_details[f"{col}{row_num}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row_num += 1

ws_details.column_dimensions['A'].width = 20
ws_details.column_dimensions['B'].width = 20
ws_details.column_dimensions['C'].width = 35
ws_details.column_dimensions['D'].width = 12
ws_details.column_dimensions['E'].width = 25
ws_details.column_dimensions['F'].width = 25
ws_details.column_dimensions['G'].width = 15
ws_details.column_dimensions['H'].width = 12

# ============================================================================
# SHEET 4: Reference
# ============================================================================

ws_ref = wb.create_sheet("Reference", 3)

ws_ref.append(["MQTT Broker", "18.143.215.113"])
ws_ref.append(["Port", "1883"])
ws_ref.append(["Protocol", "MQTTv3.1.1"])
ws_ref.append(["QoS", "1"])
ws_ref.append(["Timeout per Test", "5 seconds"])
ws_ref.append(["Total Modules", "16"])
ws_ref.append(["Total Tests", "80"])
ws_ref.append(["Test Categories", "6"])
ws_ref.append(["Date Created", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

row_num = 2
while row_num < ws_ref.max_row + 1:
    ws_ref[f"A{row_num}"].font = Font(bold=True)
    ws_ref[f"A{row_num}"].fill = category_fill
    row_num += 1

ws_ref.column_dimensions['A'].width = 20
ws_ref.column_dimensions['B'].width = 30

# Save workbook
wb.save('/home/wedman/Documents/Development/Next-NodeApps/Middleware_Testing.xlsx')

print("✅ XLSX file created successfully: Middleware_Testing.xlsx")
print("\n📊 Summary:")
print("   - Sheet 1: Middleware Test Results (16 modules × 5 tests)")
print("   - Sheet 2: Summary by Category")
print("   - Sheet 3: Test Details")
print("   - Sheet 4: Reference (Broker config, etc.)")
print("\n📍 Location: /home/wedman/Documents/Development/Next-NodeApps/Middleware_Testing.xlsx")
